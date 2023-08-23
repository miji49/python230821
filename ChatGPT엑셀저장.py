import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 엑셀 파일 생성
workbook = Workbook()
sheet = workbook.active
sheet.title = "Sales Data"

# 컬럼 헤더 생성
headers = ["제품명", "가격", "수량", "판매날짜"]
for col_idx, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)

# 데이터 생성 및 저장
for row in range(2, 102):  # 100개의 데이터 생성
    product_name = f"제품{row - 1}"
    price = round(random.uniform(100, 1000), 2)
    quantity = random.randint(1, 50)
    sale_date = f"2023-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}"

    data = [product_name, price, quantity, sale_date]
    for col_idx, value in enumerate(data, start=1):
        cell = sheet.cell(row=row, column=col_idx)
        cell.value = value

# 엑셀 파일 저장
file_path = r"c:\work\sale.xlsx"
workbook.save(file_path)
print(f"엑셀 파일이 {file_path}에 저장되었습니다.")
